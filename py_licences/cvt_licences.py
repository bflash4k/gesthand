'''
'''
import csv 
import sys
import argparse
import time
import pprint
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
 
g_lic_valide = []
g_lic_en_cours = []
g_lic_qualifie = []
g_lic_finalise = []

reload(sys)
sys.setdefaultencoding('utf-8')


def ws_append(ws, lines):
  ws.append(['',lines['\xef\xbb\xbfNom'], lines['Prenom'], lines['Sexe'], lines['N\xc3\xa9(e) le'], lines['Taille'], lines['Lateralite'], 
      lines['Numero Licence'], lines['Qualite'], lines['Nature'], lines['Ass'], lines['Type demande'], lines['Etat'], 
      lines['Num Club'], lines['Structure'], lines['Date'], lines['Transfert'], lines['Nationalite'], lines['DroitImage'], 
      lines['Utilisateur'], lines['Pole'], lines['Doubler']])


def addHeader(ws):
     ws.append(['Comment','Nom', 'Prenom', 'Sexe', 'Ne(e) le', 'Taille', 'Lateralite', 
      'Numero Licence', 'Qualite', 'Nature', 'Ass', 'Type demande', 'Etat', 
      'Num Club', 'Structure', 'Date', 'Transfert', 'Nationalite', 'DroitImage', 
      'Utilisateur', 'Pole', 'Doubler'])

def SearchAndUpdate(ws, cell, nom, prenom, naissance):
    found = 0
    for s_row in ws.iter_rows(min_row=2, max_col=5):
      for s_cell in s_row:
        s_nom = ws.cell(row=s_cell.row, column=s_cell.col_idx+1).value
        s_prenom = ws.cell(row=s_cell.row, column=s_cell.col_idx+2).value
        s_naissance = ws.cell(row=s_cell.row, column=s_cell.col_idx+4).value
        if ((s_nom == nom) and (s_prenom == prenom) and (s_naissance == naissance)):
          ##print("Found for {0} {1} -> Add comment: {2}".format(nom, prenom,cell.value))
          ws.cell(row=s_cell.row, column=s_cell.col_idx).value = cell.value
          found = 1
          break
    if (found == 0):
      print("!!!Not Found for {0} {1} -> status changed ???({2})".format(nom, prenom,cell.value))



def UpdateNewSheet(prev_file,wb):
  '''
  '''
  #Open previous file
  wb_old = load_workbook(filename=prev_file)

  #

  for sheet in wb:
    print("==================================")
    print("Process sheet {0}".format(sheet.title))
    ws = wb[sheet.title]  
    ws_old = wb_old[sheet.title]

    #for row in range(1,101): #ws_old_all.rows:
    for row in ws_old.iter_rows(min_row=2, max_col=1):
      for cell in row:
        if (cell.value):
          #print(cell.value+"+"+ cell.row +"+"+cell.col_idx)
          nom = ws_old.cell(row=cell.row, column=cell.col_idx+1).value
          prenom = ws_old.cell(row=cell.row, column=cell.col_idx+2).value
          naissance = ws_old.cell(row=cell.row, column=cell.col_idx+4).value
          #print (nom+" "+prenom+" "+naissance+":"+cell.value)

          SearchAndUpdate(ws, cell, nom, prenom, naissance)


def  OpenCSV(prev_file, gh_file):
  '''
  Open the CSV and create the workbook
  '''
  timestr = time.strftime("%H_%M_%S")
  xlsx_file = "{0}_{1}.xlsx".format(gh_file,timestr)
  print(xlsx_file)

  # Prepare the xlsx file
  wb = Workbook()
  
  ws_all = wb.active
  ws_all.title = "ALL"
  addHeader(ws_all)
  ws_qualifie = wb.create_sheet("QUALIFIE")
  addHeader(ws_qualifie)
  ws_en_cours = wb.create_sheet("EN_COURS")
  addHeader(ws_en_cours)
  ws_finalise = wb.create_sheet("FINALISE")
  addHeader(ws_finalise)
  ws_valide = wb.create_sheet("VALIDE")
  addHeader(ws_valide)

  #Process the CSV
  with open(gh_file, 'rb') as g_file:

    r = csv.DictReader(g_file, delimiter=';', quotechar='"')
    #print(r.fieldnames)
    for lines in r:
      #print(lines['\xef\xbb\xbfNom'],lines['Prenom'], lines['Etat'])
      #print(lines)

      ws_append(ws_all, lines)

      if (lines['Etat'] == 'QUALIFIE'):
        g_lic_qualifie.append(lines)
        ws_append(ws_qualifie, lines)

      elif (lines['Etat'] == 'EN_COURS'):
        g_lic_en_cours.append(lines)
        ws_append(ws_en_cours, lines)

      elif (lines['Etat'] == 'FINALISE'):
        g_lic_finalise.append(lines)
        ws_append(ws_finalise, lines)

      elif (lines['Etat'] == 'VALIDE'):
        g_lic_valide.append(lines)
        ws_append(ws_valide, lines)

      else:
        print("ERREUR ETAT INCONNU")  

  # Fix newly created xlsx by adding comments from previous/old file
  UpdateNewSheet(prev_file,wb)

  print ("save {0}".format(xlsx_file))
  wb.save(xlsx_file)

def main():
  parser = argparse.ArgumentParser(description='Parametres')
  parser.add_argument("-g","--gesthand_csv", help="Gesthand Licence CSV file", required=True)
  parser.add_argument("-p","--precedent_fichier", help="Gesthand Licence CSV file", required=False)


  args = parser.parse_args()
  OpenCSV(args.precedent_fichier, args.gesthand_csv)

if __name__ == '__main__':
  main()
